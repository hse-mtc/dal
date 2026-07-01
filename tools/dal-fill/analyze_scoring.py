"""
Сверка логики подсчёта баллов ФИЗО между:
  (A) DAL scoring.py  — strict (value <= threshold, без округления)
  (B) DAL scoring.py  — с округлением беговых результатов до 0.1 сек
  (C) ведомость 2026 "Результаты ФИЗО" — уже посчитанные баллы (столбцы «О», итог100)

Цель: понять корень расхождения и его масштаб по всем ~955 участникам.

Запуск:  python3 tools/dal-fill/analyze_scoring.py
"""
import math
import sys
from pathlib import Path

import openpyxl

# --- импорт настоящего scoring.py из back-end (чистый модуль, без Django) ---
PHYS = Path(__file__).resolve().parents[2] / "back-end" / "src" / "ams" / "physical"
sys.path.insert(0, str(PHYS))
import scoring  # noqa: E402

DATA = Path.home() / "Downloads" / "dal-fill-db"
FIZO = DATA / "Результаты ФИЗО нормативы16.06.26-2 (2).xlsx"

# Колонки 2026 "Результаты" (1-based): (result_col, dal_convert, direction, number, kind)
# kind: 'int' счётное, 'sec' секунды(бег кор.), 'mmss' мин.сек(бег длин.), 'kettle' гиря
STRENGTH, SPEED, ENDUR = "ST", "SP", "EN"
EXODEF = [
    # подтягивание №3
    ("PULL_UPS", 12, 13, scoring.convert_pull_ups, STRENGTH, 3, "int"),
    # подъём переворотом №5
    ("LIFT_TURNOVER", 14, 15, scoring.convert_lift_turnover, STRENGTH, 5, "int"),
    # подъём силой №6
    ("LIFT_FORCE", 16, 17, scoring.convert_lift_force, STRENGTH, 6, "int"),
    # рывок гири №10 — результат в R/S/T (70/80/80+), балл в U(21)
    ("KETTLEBELL_SNATCH", None, 21, scoring.convert_kettlebell_snatch, STRENGTH, 10, "kettle"),
    # бег 60м №17
    ("RUN_60", 23, 24, scoring.convert_run_60m, SPEED, 17, "sec"),
    # бег 100м №18
    ("RUN_100", 25, 26, scoring.convert_run_100m, SPEED, 18, "sec"),
    # челночный 10x10 №19
    ("SHUTTLE_RUN", 27, 28, scoring.convert_shuttle_run, SPEED, 19, "sec"),
    # бег 1км №24
    ("RUN_1K", 30, 31, scoring.convert_run_1km, ENDUR, 24, "mmss"),
    # бег 3км №25
    ("RUN_3K", 32, 33, scoring.convert_run_3km, ENDUR, 25, "mmss"),
]
# столбцы гири по весовым категориям (result → вес для DAL)
KETTLE_COLS = {18: 65, 19: 75, 20: 85}  # R=<70(light), S=70-80(medium), T=>80(heavy)

COL_SILA_O, COL_BYSTR_O, COL_VYNOS_O = 22, 29, 34
COL_ITOG, COL_OC5, COL_ITOG100 = 35, 36, 37


def round_01(x):
    """Округление до 0.1 (half-up) — как в формулах ведомости."""
    return math.floor(x * 10 + 0.5) / 10.0


def mmss_to_decmin(v):
    """3.51 (3 мин 51 сек) -> 3.85 десятичных минут."""
    minutes = int(v)
    seconds = round((v - minutes) * 100)  # .51 -> 51 сек
    return minutes + seconds / 60.0


def age_of(bd):
    from datetime import date
    today = date(2026, 7, 2)
    return today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))


def num(v):
    if v in (None, ""):
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def main():
    wb = openpyxl.load_workbook(FIZO, data_only=True)
    ws = wb["до 25"]

    rows = []
    for r in range(15, 970):
        name = ws.cell(r, 2).value
        if not name:
            continue
        bd = ws.cell(r, 3).value
        rows.append((r, name, bd))

    # счётчики
    n = 0
    per_ex_strict_mismatch = {d[0]: 0 for d in EXODEF}
    per_ex_round_mismatch = {d[0]: 0 for d in EXODEF}
    per_ex_present = {d[0]: 0 for d in EXODEF}
    grade_strict_mismatch = 0
    grade_round_mismatch = 0
    examples_strict = []
    examples_round = []

    for (r, name, bd) in rows:
        n += 1
        age = age_of(bd) if bd else 19
        dir_scores_strict = {STRENGTH: [], SPEED: [], ENDUR: []}
        dir_scores_round = {STRENGTH: [], SPEED: [], ENDUR: []}

        for (etype, rcol, ocol, conv, direction, number, kind) in EXODEF:
            # получить сырой результат и балл ведомости
            ved_o = num(ws.cell(r, ocol).value)
            if kind == "kettle":
                val = None
                weight = None
                for kcol, w in KETTLE_COLS.items():
                    cv = num(ws.cell(r, kcol).value)
                    if cv > 0:
                        val, weight = cv, w
                        break
                if val is None:
                    continue
                per_ex_present[etype] += 1
                s_strict = conv(val, {"weight": weight}, age)
                s_round = s_strict  # счётное, округление не нужно
            else:
                raw = num(ws.cell(r, rcol).value)
                if raw <= 0:
                    continue
                per_ex_present[etype] += 1
                if kind == "mmss":
                    dv = mmss_to_decmin(raw)
                    s_strict = conv(dv, {}, age)
                    s_round = conv(dv, {}, age)  # секунды точны -> без округл.
                elif kind == "sec":
                    s_strict = conv(raw, {}, age)
                    s_round = conv(round_01(raw), {}, age)
                else:  # int
                    s_strict = conv(int(raw), {}, age)
                    s_round = s_strict

            if s_strict != ved_o:
                per_ex_strict_mismatch[etype] += 1
                if len(examples_strict) < 25:
                    examples_strict.append(
                        (name, etype, ws.cell(r, rcol).value if rcol else "kettle",
                         f"DAL={s_strict}", f"вед={ved_o}"))
            if s_round != ved_o:
                per_ex_round_mismatch[etype] += 1
                if len(examples_round) < 25:
                    examples_round.append(
                        (name, etype, ws.cell(r, rcol).value if rcol else "kettle",
                         f"DALr={s_round}", f"вед={ved_o}"))

            dir_scores_strict[direction].append(s_strict)
            dir_scores_round[direction].append(s_round)

        # итоговый балл (100-балльный global)
        def grade(dsc):
            st = max(dsc[STRENGTH]) if dsc[STRENGTH] else 0
            sp = max(dsc[SPEED]) if dsc[SPEED] else 0
            en = max(dsc[ENDUR]) if dsc[ENDUR] else 0
            return scoring.compute_global_score(st, sp, en, age)

        g_strict = grade(dir_scores_strict)
        g_round = grade(dir_scores_round)
        ved_grade = int(num(ws.cell(r, COL_ITOG100).value))

        if g_strict != ved_grade:
            grade_strict_mismatch += 1
        if g_round != ved_grade:
            grade_round_mismatch += 1

    print(f"Всего участников: {n}\n")
    print("Наличие результатов по упражнениям:")
    for d in EXODEF:
        print(f"  {d[0]:16s} present={per_ex_present[d[0]]:4d}")
    print("\nРасхождения ПОБАЛЛЬНО (secondary score) DAL vs ведомость:")
    print(f"  {'упражнение':16s} {'strict':>8s} {'round0.1':>9s}")
    for d in EXODEF:
        print(f"  {d[0]:16s} {per_ex_strict_mismatch[d[0]]:8d} {per_ex_round_mismatch[d[0]]:9d}")
    print("\nРасхождения ИТОГОВОГО балла (100-балльн.) vs ведомость (столбец итог100):")
    print(f"  strict  : {grade_strict_mismatch} / {n}")
    print(f"  round0.1: {grade_round_mismatch} / {n}")

    print("\nПримеры расхождений STRICT (первые):")
    for e in examples_strict[:15]:
        print("   ", e)
    print("\nПримеры расхождений ROUND0.1 (первые):")
    for e in examples_round[:15]:
        print("   ", e)


if __name__ == "__main__":
    main()


# ============================================================================
# ДОП. ВЕРИФИКАЦИЯ: (1) таблицы DAL == сетка комиссии; (2) grade по О комиссии
# == итог100 комиссии; (3) «примагничивание» value к сетке -> DAL == комиссия.
# ============================================================================

def verify_gridsnap():
    import openpyxl as _oxl
    wb = _oxl.load_workbook(FIZO, data_only=True)
    ws = wb["до 25"]

    # инверсия: для «меньше-лучше» вернуть порог-сетку, дающий score S под DAL.
    def invert_time(table, S):
        # table: [(threshold_time, score)] возрастающее время / убывающий балл
        cands = [t for (t, sc) in table if sc == S]
        return max(cands) if cands else None

    TIME_TABLES = {
        "RUN_60": scoring._RUN_60M_TABLE,
        "RUN_100": scoring._RUN_100M_TABLE,
        "SHUTTLE_RUN": scoring._SHUTTLE_RUN_TABLE,
        "RUN_1K": scoring._RUN_1KM_TABLE,
        "RUN_3K": scoring._RUN_3KM_TABLE,
    }

    n = 0
    grade_by_comm_o_mismatch = 0     # grade из О комиссии vs итог100 комиссии
    grade_snap_mismatch = 0          # grade из snapped-value(DAL) vs итог100 комиссии
    unreproducible = []              # О комиссии не выражается порогом DAL
    o_table_mismatch = []            # О комиссии не совпадает с DAL(snap)

    for r in range(15, 970):
        name = ws.cell(r, 2).value
        if not name:
            continue
        bd = ws.cell(r, 3).value
        age = age_of(bd) if bd else 19
        n += 1

        comm_dir = {STRENGTH: [], SPEED: [], ENDUR: []}
        snap_dir = {STRENGTH: [], SPEED: [], ENDUR: []}

        for (etype, rcol, ocol, conv, direction, number, kind) in EXODEF:
            comm_o = int(num(ws.cell(r, ocol).value))
            if kind == "kettle":
                present = any(num(ws.cell(r, c).value) > 0 for c in KETTLE_COLS)
                if not present:
                    continue
                # для гири храним сырое: DAL уже совпал 100%
                snap_o = comm_o
            else:
                raw = num(ws.cell(r, rcol).value)
                if raw <= 0:
                    continue
                if kind == "int":
                    snap_o = comm_o  # счётные совпали 100%
                else:
                    tbl = TIME_TABLES[etype]
                    snap_val = invert_time(tbl, comm_o)
                    if snap_val is None:
                        unreproducible.append((name, etype, comm_o))
                        snap_o = None
                    else:
                        snap_o = conv(snap_val, {}, age)
                        if snap_o != comm_o:
                            o_table_mismatch.append((name, etype, comm_o, snap_o))
            if snap_o is not None:
                snap_dir[direction].append(snap_o)
            comm_dir[direction].append(comm_o)

        def g(dsc):
            st = max(dsc[STRENGTH]) if dsc[STRENGTH] else 0
            sp = max(dsc[SPEED]) if dsc[SPEED] else 0
            en = max(dsc[ENDUR]) if dsc[ENDUR] else 0
            return scoring.compute_global_score(st, sp, en, age)

        ved_grade = int(num(ws.cell(r, COL_ITOG100).value))
        if g(comm_dir) != ved_grade:
            grade_by_comm_o_mismatch += 1
        if g(snap_dir) != ved_grade:
            grade_snap_mismatch += 1

    print("\n" + "=" * 70)
    print("ВЕРИФИКАЦИЯ GRID-SNAP (хранить значение, примагниченное к сетке):")
    print(f"  участников: {n}")
    print(f"  grade(из О комиссии)      != итог100 комиссии : {grade_by_comm_o_mismatch}")
    print(f"  grade(DAL по snapped val) != итог100 комиссии : {grade_snap_mismatch}")
    print(f"  О комиссии невыразимо порогом DAL : {len(unreproducible)}")
    if unreproducible[:10]:
        print("    примеры:", unreproducible[:10])
    print(f"  DAL(snap) != О комиссии (таблицы расходятся) : {len(o_table_mismatch)}")
    if o_table_mismatch[:10]:
        print("    примеры:", o_table_mismatch[:10])


if __name__ == "__main__":
    verify_gridsnap()
