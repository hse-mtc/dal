#!/usr/bin/env python3
"""
reconcile.py — сверка НАМЕРЕННЫХ (intended) значений протокола для ДВУХ путей
заполнения против единого источника истины core.compute_fill.

Идея: и Вариант 1 (ручной xlsx, fill_manual.py), и Вариант 2 (prod-план,
fill_prod.py) обязаны нести ОДНИ И ТЕ ЖЕ числа, потому что оба берут их из
core.compute_fill. Этот скрипт независимо пересчитывает core.compute_fill для
КАЖДОГО человека, присутствующего в ведомости ФИЗО (Москва+ВАВТ), и для каждого
человека со средним баллом, и проверяет:

  (a) Значения, которые fill_manual РЕАЛЬНО записал в artifacts/manual/*.xlsx
      (переоткрываем xlsx, читаем колонки P(16)/R(18)/U(21)/V(22), сопоставляем
      по имени+дате), РАВНЫ выходу core.compute_fill.
      Расхождений быть не должно (0).

  (b) План artifacts/prod/plan.json кодирует ТЕ ЖЕ значения упражнений
      (value = value_to_store) и тот же средний балл, что даёт core.compute_fill
      (значит DAL, пересчитав из этих значений, придёт к тому же протоколу).

Точка входа сравнения — зеркало логики записи fill_manual._fill_sheet:
  * matched (ФИЗО найден) -> P = total_score, R = physical_test_grade;
  * grade найден         -> U = mean_grade_scaled;
  * V = final_result всегда.
Ячейки, которые fill_manual НЕ пишет, мы НЕ сверяем (они сохраняют исходное
значение шаблона и к «намеренным» значениям отношения не имеют).

Запуск:  python3 tools/dal-fill/reconcile.py
Итоговый stdout = возвращаемое значение (машиночитаемый JSON-хвост + сводка).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import core  # noqa: E402

# ---------------------------------------------------------------------------
# Пути к артефактам обоих вариантов.
# ---------------------------------------------------------------------------
MANUAL_DIR = core.DATA / "artifacts" / "manual"
PLAN_PATH = core.DATA / "artifacts" / "prod" / "plan.json"

# Колонки протокола, которые fill_manual заполняет (1-based), зеркало
# fill_manual.COL. Сверяем только их.
COL_TOTAL = 16       # P  = total_score          (только если matched)
COL_PHYS_GRADE = 18  # R  = physical_test_grade   (только если matched)
COL_MEAN_SCALED = 21  # U = mean_grade_scaled     (только если grade найден)
COL_FINAL = 22       # V  = final_result          (всегда)

DATA_START_ROW = 20


# ===========================================================================
# Индекс строк протоколов из artifacts/manual/*.xlsx
# ===========================================================================
def build_manual_index(manual_dir: Path):
    """Собрать (norm_name, norm_dob) -> {file, sheet, row, P, R, U, V}.

    Идём по всем xlsx, всем листам, строкам данных с DATA_START_ROW до начала
    раздела 2 (не прошедшие отбор) — ровно как fill_manual.fill_sheet.
    Разбор ячейки B через ту же логику 'ФИО\\nDD.MM.YYYY'.
    """
    index: dict[tuple[str, str], dict] = {}
    files = sorted(
        p for p in manual_dir.glob("*.xlsx") if not p.name.startswith("~$")
    )
    for f in files:
        wb = openpyxl.load_workbook(f)
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r in range(DATA_START_ROW, ws.max_row + 1):
                a = ws.cell(r, 1).value
                if isinstance(a, str) and a.strip().startswith("2. Список"):
                    break
                b = ws.cell(r, 2).value
                if not b or "\n" not in str(b):
                    continue
                parts = str(b).split("\n")
                name = parts[0].strip()
                dob = parts[1].strip() if len(parts) > 1 else ""
                if not name:
                    continue
                key = core._key(name, dob)
                # На случай дубликатов ключа — берём ПЕРВЫЙ (как и остальные
                # индексы в проекте); дубликаты фиксируем отдельно.
                if key in index:
                    index[key].setdefault("_dups", []).append(
                        {"file": f.name, "sheet": sn, "row": r}
                    )
                    continue
                index[key] = {
                    "file": f.name,
                    "sheet": sn,
                    "row": r,
                    "name": name,
                    "dob": dob,
                    "P": ws.cell(r, COL_TOTAL).value,
                    "R": ws.cell(r, COL_PHYS_GRADE).value,
                    "U": ws.cell(r, COL_MEAN_SCALED).value,
                    "V": ws.cell(r, COL_FINAL).value,
                }
    return index, [f.name for f in files]


def _as_num(v):
    """Толерантно привести значение ячейки к числу для сравнения."""
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return v


# ===========================================================================
# Набор людей для сверки: все ФИЗО (Москва+ВАВТ) + все со средним баллом.
# ===========================================================================
def build_people(fizo: core.FizoIndex, grades: core.GradeIndex):
    """Вернуть список (name, dob_raw, sources) уникальных людей.

    sources: множество из {'fizo','grade'}. Ключ уникальности — норм.(name,dob).
    Кампусный охват физподготовки (Москва+ВАВТ) обеспечен самим источником ФИЗО
    — в ведомости нет других кампусов. Средний балл — кто угодно.
    """
    people: dict[tuple[str, str], dict] = {}
    for rec in fizo.records:
        key = core._key(rec["name"], rec["dob_raw"])
        p = people.setdefault(
            key, {"name": rec["name"], "dob_raw": rec["dob_raw"], "sources": set()}
        )
        p["sources"].add("fizo")
    for rec in grades.records:
        key = core._key(rec["name"], rec["dob_raw"])
        p = people.setdefault(
            key, {"name": rec["name"], "dob_raw": rec["dob_raw"], "sources": set()}
        )
        p["sources"].add("grade")
    return people


# ===========================================================================
# Ожидаемые значения ячеек манульного протокола — зеркало fill_manual.
# ===========================================================================
def expected_manual_cells(fill: dict) -> dict:
    """Что fill_manual ЗАПИСАЛ БЫ для данного человека.

    Возвращает dict только с теми колонками, которые fill_manual реально пишет:
      matched -> P (total_score), R (physical_test_grade)
      mean_grade найден -> U (mean_grade_scaled)
      всегда -> V (final_result)
    """
    exp: dict[str, object] = {}
    if fill["matched"]:
        exp["P"] = fill["total_score"]
        exp["R"] = fill["physical_test_grade"]
    if fill["mean_grade_scaled"] is not None:
        exp["U"] = fill["mean_grade_scaled"]
    exp["V"] = fill["final_result"]
    return exp


# ===========================================================================
# (a) Сверка manual xlsx
# ===========================================================================
def check_manual(people, fizo, grades, manual_index):
    """Вернуть (checked, mismatches[list])."""
    checked = 0
    mismatches = []
    not_in_protocol = []  # человек есть в источнике, но строки в протоколе нет

    for key, person in people.items():
        name = person["name"]
        dob = person["dob_raw"]
        fill = core.compute_fill(name, dob, fizo, grades)
        row = manual_index.get(key)
        if row is None:
            not_in_protocol.append(
                {"name": name, "dob": core.norm_dob(dob),
                 "sources": sorted(person["sources"])}
            )
            continue

        checked += 1
        exp = expected_manual_cells(fill)
        for col_letter, cell_key in (("P", "P"), ("R", "R"), ("U", "U"), ("V", "V")):
            if col_letter not in exp:
                continue  # эту колонку fill_manual не пишет — не сверяем
            want = _as_num(exp[col_letter])
            got = _as_num(row[cell_key])
            if want != got:
                mismatches.append({
                    "name": name,
                    "dob": core.norm_dob(dob),
                    "file": row["file"],
                    "sheet": row["sheet"],
                    "row": row["row"],
                    "column": col_letter,
                    "expected": want,
                    "actual": got,
                    "matched_physical": fill["matched"],
                    "has_grade": fill["mean_grade_scaled"] is not None,
                })
    return checked, mismatches, not_in_protocol


# ===========================================================================
# (b) Сверка prod plan.json
# ===========================================================================
def _plan_exercise_index(plan: dict):
    """(norm_name, norm_dob, exercise_type) -> {value, secondary_score, extra_params}.

    Читает новую структуру physical_overrides[].body.results (ручка override).
    """
    idx = {}
    for item in plan.get("physical_overrides", []):
        nn, nd = core.norm_name(item["name"]), core.norm_dob(item["dob"])
        for er in item.get("body", {}).get("results", []):
            idx[(nn, nd, er["exercise_type"])] = er
    return idx


def _plan_grade_index(plan: dict):
    """(norm_name, norm_dob) -> committee_mean_grade (из grade_patches)."""
    idx = {}
    for item in plan.get("grade_patches", []):
        k = (core.norm_name(item["name"]), core.norm_dob(item["dob"]))
        idx[k] = item["committee_mean_grade"]
    return idx


def check_plan(people, fizo, grades, plan):
    """Сверить, что план кодирует ТЕ ЖЕ per-person значения упражнений и средний
    балл, что даёт core.compute_fill.

    Проверяем ДВА направления:
      1. Каждое упражнение из core.compute_fill.exercise_results для человека,
         КОТОРЫЙ ЕСТЬ в карте плана, присутствует в плане с тем же value и
         extra_params (иначе DAL пересчитает иначе).
      2. Средний балл: для человека в grade_patches плана committee_mean_grade
         == core mean_grade. (grade_conflicts плана НЕ трогают mean_grade, и
         совпавшие с БД тоже — для мок-плана этого не бывает.)

    Возвращает (checked, mismatches[list]).
    """
    ex_idx = _plan_exercise_index(plan)
    grade_idx = _plan_grade_index(plan)

    # Множество (name,dob), которых план вообще касается (touched) — по факту
    # наличия хоть в одном ex/grade-плане. Только их значения план кодирует.
    plan_people = set()
    for (nn, nd, _et) in ex_idx:
        plan_people.add((nn, nd))
    for (nn, nd) in grade_idx:
        plan_people.add((nn, nd))

    checked = 0
    mismatches = []

    for key, person in people.items():
        name = person["name"]
        dob = person["dob_raw"]
        pkey = (core.norm_name(name), core.norm_dob(dob))
        if pkey not in plan_people:
            # План этого человека не касается (например, graded-only, которого
            # нет в ФИЗО-мок-карте плана). Нечего сверять — не расхождение.
            continue

        checked += 1
        fill = core.compute_fill(name, dob, fizo, grades)

        # --- 1. упражнения ---
        for er in fill["exercise_results"]:
            etype = er["exercise_type"]
            want_value = _as_num(er["value_to_store"])
            want_extra = er["extra_params"] or {}
            planned = ex_idx.get((pkey[0], pkey[1], etype))
            if planned is None:
                mismatches.append({
                    "kind": "exercise_missing_in_plan",
                    "name": name, "dob": core.norm_dob(dob),
                    "exercise_type": etype,
                    "expected_value": want_value,
                })
                continue
            got_value = _as_num(planned.get("value"))
            got_extra = planned.get("extra_params") or {}
            if want_value != got_value:
                mismatches.append({
                    "kind": "exercise_value",
                    "name": name, "dob": core.norm_dob(dob),
                    "exercise_type": etype,
                    "expected_value": want_value, "actual_value": got_value,
                })
            # Дословный балл комиссии (verbatim): план обязан нести тот же
            # secondary_score, что и core (иначе протокол разойдётся).
            want_score = er.get("secondary_score")
            got_score = planned.get("secondary_score")
            if want_score is not None and want_score != got_score:
                mismatches.append({
                    "kind": "exercise_secondary_score",
                    "name": name, "dob": core.norm_dob(dob),
                    "exercise_type": etype,
                    "expected_score": want_score, "actual_score": got_score,
                })
            if want_extra != got_extra:
                mismatches.append({
                    "kind": "exercise_extra_params",
                    "name": name, "dob": core.norm_dob(dob),
                    "exercise_type": etype,
                    "expected_extra": want_extra, "actual_extra": got_extra,
                })

        # Обратная сторона: план не должен содержать «лишних» упражнений для
        # этого человека, которых нет в core.exercise_results.
        core_etypes = {er["exercise_type"] for er in fill["exercise_results"]}
        for (nn, nd, et) in ex_idx:
            if (nn, nd) == pkey and et not in core_etypes:
                mismatches.append({
                    "kind": "extra_exercise_in_plan",
                    "name": name, "dob": core.norm_dob(dob),
                    "exercise_type": et,
                    "actual_value": _as_num(ex_idx[(nn, nd, et)].get("value")),
                })

        # --- 2. средний балл ---
        planned_grade = grade_idx.get(pkey)
        if planned_grade is not None:
            want_grade = fill["mean_grade"]
            if want_grade is None or abs(float(planned_grade) - float(want_grade)) > 1e-9:
                mismatches.append({
                    "kind": "mean_grade",
                    "name": name, "dob": core.norm_dob(dob),
                    "expected_mean_grade": want_grade,
                    "actual_mean_grade": planned_grade,
                })

    return checked, mismatches


# ===========================================================================
# main
# ===========================================================================
def main() -> int:
    fizo = core.load_fizo()
    grades = core.load_grades()
    people = build_people(fizo, grades)

    manual_index, manual_files = build_manual_index(MANUAL_DIR)
    plan = json.loads(PLAN_PATH.read_text(encoding="utf-8"))

    checked_a, mism_a, not_in_proto = check_manual(
        people, fizo, grades, manual_index
    )
    checked_b, mism_b = check_plan(people, fizo, grades, plan)

    print("=" * 72)
    print("RECONCILE — intended protocol values vs core.compute_fill")
    print("=" * 72)
    print(f"SCORING_MODE           : {core.SCORING_MODE}")
    print(f"ФИЗО people (unique)   : {len({core._key(r['name'], r['dob_raw']) for r in fizo.records})}")
    print(f"graded people (unique) : {len({core._key(r['name'], r['dob_raw']) for r in grades.records})}")
    print(f"people to reconcile    : {len(people)}  (union ФИЗО ∪ graded)")
    print(f"manual protocol files  : {manual_files}")
    print(f"plan applicants_in_map : {plan['meta']['applicants_in_map']}")
    print("-" * 72)
    print("(a) MANUAL xlsx  (cols P/R/U/V by name+dob) vs core.compute_fill")
    print(f"    checked (rows found)     : {checked_a}")
    print(f"    people not in any proto  : {len(not_in_proto)} (cannot check — no row)")
    print(f"    MISMATCHES               : {len(mism_a)}")
    print("-" * 72)
    print("(b) PROD plan.json (per-person exercise value + mean_grade) vs core")
    print(f"    checked (plan-covered)   : {checked_b}")
    print(f"    MISMATCHES               : {len(mism_b)}")
    print("=" * 72)

    if mism_a:
        print("\n(a) up to 10 example mismatches:")
        for m in mism_a[:10]:
            print("   ", json.dumps(m, ensure_ascii=False, default=str))
    if mism_b:
        print("\n(b) up to 10 example mismatches:")
        for m in mism_b[:10]:
            print("   ", json.dumps(m, ensure_ascii=False, default=str))

    summary = {
        "scoring_mode": core.SCORING_MODE,
        "people_reconciled": len(people),
        "a_manual": {
            "checked": checked_a,
            "mismatches": len(mism_a),
            "people_not_in_protocol": len(not_in_proto),
            "examples": mism_a[:10],
        },
        "b_plan": {
            "checked": checked_b,
            "mismatches": len(mism_b),
            "examples": mism_b[:10],
        },
        "ok": len(mism_a) == 0 and len(mism_b) == 0,
    }
    print("\nSUMMARY_JSON " + json.dumps(summary, ensure_ascii=False, default=str))
    return 0 if summary["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
