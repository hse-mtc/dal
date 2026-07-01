"""
core.py — единый общий модуль для ДВУХ путей заполнения "Протокола конкурсного
отбора" физическими результатами (ФИЗО) и средним баллом:

  * Вариант 1 (ручной): пишем значения прямо в копии существующих xlsx-протоколов.
  * Вариант 2 (prod):    пушим сырые значения в prod-API DAL, DAL сам пересчитывает
                          и заново экспортирует протокол.

Оба варианта импортируют ЭТОТ модуль и берут из него ОДНИ И ТЕ ЖЕ числа
(value_to_store, score, physical_test_grade, mean_grade_scaled, final_result),
поэтому итоговые протоколы обязаны совпасть побайтно по значениям.

Проверенная логика (из tools/dal-fill/analyze_scoring.py) переиспользуется здесь
без изменений: карта колонок упражнений (EXODEF), mmss_to_decmin(), round_01(),
трюк импорта настоящего scoring.py и инверсия «примагничивания» к сетке
(invert_time). Прогон analyze_scoring доказал: если хранить примагниченное к
сетке значение, DAL воспроизводит баллы комиссии (0/955 расхождений).

Зеркалит back-end/src/ams/utils/export/comp_sel_protocol.py:
  _select_best_exercises (лучшее в направлении = MAX score), EXERCISE_NUMBERS,
  _format_exercise_value, _get_age_group.
Опорная дата возраста — 2026-07-02.

Запуск:  python3 tools/dal-fill/core.py
"""

from __future__ import annotations

import math
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Импорт настоящего scoring.py из back-end (чистый модуль, без Django deps)
# ---------------------------------------------------------------------------
PHYS = Path(__file__).resolve().parents[2] / "back-end" / "src" / "ams" / "physical"
sys.path.insert(0, str(PHYS))
import scoring  # noqa: E402

# ---------------------------------------------------------------------------
# Режим подсчёта
#   'verbatim' (по умолчанию): храним СЫРОЙ результат комиссии (спринт 8.33,
#       длинные — в десятичных минутах 3.85 = 3:51) и ДОСЛОВНЫЙ балл комиссии О.
#       Заливается через ручку override (secondary_score пишется как есть, без
#       пересчёта). В протоколе — сырой результат + мм:сс + балл комиссии.
#   'committee': примагничиваем время к сетке DAL так, чтобы DAL сам пересчитал
#       и воспроизвёл подписанный балл О (0/955 расхождений). Обычные ручки.
#   'strict': храним сырые значения, DAL считает строго value <= threshold.
# ---------------------------------------------------------------------------
SCORING_MODE = "verbatim"

# Беговые длинные дистанции показываются в протоколе как мин:сек (зеркало
# back-end/src/ams/utils/export/comp_sel_protocol._MMSS_EXERCISES).
_MMSS_EXERCISES = {"RUN_1K", "RUN_3K"}

# ---------------------------------------------------------------------------
# Пути к источникам
# ---------------------------------------------------------------------------
DATA = Path.home() / "Downloads" / "dal-fill-db"
FIZO_PATH = DATA / "Результаты ФИЗО нормативы16.06.26-2 (2).xlsx"
FIZO_SHEET = "до 25"
FIZO_ROW_START = 15
FIZO_ROW_END = 969  # включительно
FIZO_NAME_COL = 2
FIZO_DOB_COL = 3

GRADE_PATH = DATA / "Средний балл.xlsx"
GRADE_SHEET = "Лист2"
GRADE_NAME_COL = 1  # A = ФИО
GRADE_DOB_COL = 2   # B = дата рождения
GRADE_MEAN_COL = 5  # E = средний балл 0..10

# Опорная дата для расчёта возраста (совпадает с текущей датой протокола).
AGE_REF_DATE = date(2026, 7, 2)

# ---------------------------------------------------------------------------
# Направления (значения совпадают с Direction.* из ams.physical.exercises)
# ---------------------------------------------------------------------------
STRENGTH = "ST"
SPEED = "SP"
ENDURANCE = "EN"

# ---------------------------------------------------------------------------
# Определение упражнений (переиспользуем проверенную карту из analyze_scoring).
# Формат: (exercise_type, result_col, score_col, convert_fn, direction, number, kind)
#   result_col — 1-based колонка сырого результата (None для гири — особый разбор)
#   score_col  — 1-based колонка балла «О», подписанного комиссией
#   number     — номер упражнения в протоколе (EXERCISE_NUMBERS)
#   kind       — 'int' счётное | 'sec' секунды (короткий бег) |
#                'mmss' мин.сек (длинный бег) | 'kettle' гиря
# exercise_type здесь — это ExerciseType.value (то, что хранит DAL).
# ---------------------------------------------------------------------------
EXODEF = [
    ("PULL_UPS", 12, 13, scoring.convert_pull_ups, STRENGTH, 3, "int"),
    ("LIFT_TURNOVER", 14, 15, scoring.convert_lift_turnover, STRENGTH, 5, "int"),
    ("LIFT_FORCE", 16, 17, scoring.convert_lift_force, STRENGTH, 6, "int"),
    ("KETTLEBELL_SNATCH", None, 21, scoring.convert_kettlebell_snatch, STRENGTH, 10, "kettle"),
    ("RUN_60", 23, 24, scoring.convert_run_60m, SPEED, 17, "sec"),
    ("RUN_100", 25, 26, scoring.convert_run_100m, SPEED, 18, "sec"),
    ("SHUTTLE_RUN", 27, 28, scoring.convert_shuttle_run, SPEED, 19, "sec"),
    ("RUN_1K", 30, 31, scoring.convert_run_1km, ENDURANCE, 24, "mmss"),
    ("RUN_3K", 32, 33, scoring.convert_run_3km, ENDURANCE, 25, "mmss"),
]

# Колонки гири по весовым категориям (result_col -> вес атлета для DAL).
#   R(18) — до 70 кг (light, weight=65)
#   S(19) — 70-80 кг (medium, weight=75)
#   T(20) — более 80 кг (heavy, weight=85)
KETTLE_COLS = {18: 65, 19: 75, 20: 85}

# Агрегатные колонки, подписанные комиссией (1-based), для verbatim-режима:
#   СИЛА-О=22, БЫСТР-О=29, ВЫНОС-О=34, итог по 100-балльной=37.
COL_STRENGTH_O = 22
COL_SPEED_O = 29
COL_ENDURANCE_O = 34
COL_ITOG100 = 37

# Номер упражнения в протоколе — из comp_sel_protocol.EXERCISE_NUMBERS
# (уже зашит в EXODEF полем number, дублируем как справочник по exercise_type).
EXERCISE_NUMBERS = {d[0]: d[5] for d in EXODEF}

# Таблицы «меньше — лучше» для инверсии примагничивания к сетке.
TIME_TABLES = {
    "RUN_60": scoring._RUN_60M_TABLE,
    "RUN_100": scoring._RUN_100M_TABLE,
    "SHUTTLE_RUN": scoring._SHUTTLE_RUN_TABLE,
    "RUN_1K": scoring._RUN_1KM_TABLE,
    "RUN_3K": scoring._RUN_3KM_TABLE,
}


# ===========================================================================
# Числовые/строковые хелперы (переиспользованы из analyze_scoring.py)
# ===========================================================================
def round_01(x: float) -> float:
    """Округление до 0.1 (half-up) — как в формулах ведомости."""
    return math.floor(x * 10 + 0.5) / 10.0


def mmss_to_decmin(v: float) -> float:
    """3.51 (3 мин 51 сек) -> 3.85 десятичных минут (то, что ждёт DAL)."""
    minutes = int(v)
    seconds = round((v - minutes) * 100)  # .51 -> 51 сек
    return minutes + seconds / 60.0


def invert_time(table, target_score: int):
    """Для «меньше-лучше» вернуть максимальный порог сетки, дающий target_score
    под DAL (value <= threshold). Так DAL воспроизведёт балл О комиссии.

    table: [(threshold_time, score)] — время возрастает, балл убывает.
    """
    cands = [t for (t, sc) in table if sc == target_score]
    return max(cands) if cands else None


def _age_of(bd: date) -> int:
    """Полных лет на опорную дату AGE_REF_DATE."""
    return (
        AGE_REF_DATE.year
        - bd.year
        - ((AGE_REF_DATE.month, AGE_REF_DATE.day) < (bd.month, bd.day))
    )


def _num(v):
    """Толерантное приведение ячейки к float; пусто/мусор -> 0."""
    if v in (None, ""):
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


# ===========================================================================
# Нормализация ключей сопоставления (name, dob)
# ===========================================================================
def norm_name(s) -> str:
    """NFC, ё->е, схлопнуть пробелы, casefold."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    s = s.replace("ё", "е").replace("Ё", "Е")
    s = " ".join(s.split())
    return s.casefold()


def norm_dob(v) -> str:
    """Нормализовать дату рождения к 'dd.mm.yyyy'.

    Принимает datetime/date, а также строки в формах dd.mm.yyyy / yyyy-mm-dd
    / dd/mm/yyyy. Пусто/непонятно -> ''.
    """
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, date):
        return v.strftime("%d.%m.%Y")
    s = str(v).strip()
    if not s:
        return ""
    # Отрезать возможную временную часть.
    s = s.split(" ")[0].split("T")[0]
    for sep in (".", "-", "/"):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                a, b, c = parts
                # yyyy-mm-dd
                if len(a) == 4:
                    y, m, d = a, b, c
                else:  # dd.mm.yyyy / dd/mm/yyyy
                    d, m, y = a, b, c
                try:
                    return f"{int(d):02d}.{int(m):02d}.{int(y):04d}"
                except ValueError:
                    return s
    return s


def _key(name, dob) -> tuple[str, str]:
    return (norm_name(name), norm_dob(dob))


# ===========================================================================
# Загрузчики: строят индексы по нормализованному (name, dob) + name-only fallback
# ===========================================================================
class FizoIndex:
    """Индекс результатов ФИЗО.

    by_key      : (norm_name, norm_dob) -> record
    by_name     : norm_name -> [record, ...]   (для уникального fallback по имени)
    records     : список всех record (для отчётов о непопавших)

    record = {
        "name": str, "dob_raw": <cell>, "dob": "dd.mm.yyyy",
        "age": int, "row": int,
        "raw": {exercise_type: (result_col_value | ('kettle', reps, weight))},
    }
    """

    def __init__(self):
        self.by_key: dict[tuple[str, str], dict] = {}
        self.by_name: dict[str, list[dict]] = {}
        self.records: list[dict] = []

    def add(self, rec: dict) -> None:
        self.records.append(rec)
        self.by_key[_key(rec["name"], rec["dob_raw"])] = rec
        self.by_name.setdefault(norm_name(rec["name"]), []).append(rec)

    def lookup(self, name, dob):
        """Вернуть (record, match_kind) или (None, None).

        match_kind: 'key' (name+dob) | 'name' (уникальный name-only fallback).
        """
        rec = self.by_key.get(_key(name, dob))
        if rec is not None:
            return rec, "key"
        cands = self.by_name.get(norm_name(name), [])
        if len(cands) == 1:
            return cands[0], "name"
        return None, None


def load_fizo(path: Path = FIZO_PATH) -> FizoIndex:
    """Загрузить ФИЗО в индекс (единственный канонический загрузчик).

    Читает по каждому упражнению И сырой результат (rec["raw"][etype]), И
    подписанный комиссией балл О (rec["committee_o"][etype]) — последний нужен
    committee-режиму для инверсии примагничивания к сетке.

    Кампусный охват: в ведомости только Москва + ВАВТ (нет СПб/НН/Перми) —
    фильтрацию по кампусу делает вызывающий, здесь индексируем всех со строкой.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[FIZO_SHEET]

    index = FizoIndex()
    for r in range(FIZO_ROW_START, FIZO_ROW_END + 1):
        name = ws.cell(r, FIZO_NAME_COL).value
        if not name:
            continue
        dob_raw = ws.cell(r, FIZO_DOB_COL).value

        raw: dict[str, object] = {}
        committee_o: dict[str, int] = {}
        for (etype, rcol, ocol, conv, direction, number, kind) in EXODEF:
            committee_o[etype] = int(_num(ws.cell(r, ocol).value))
            if kind == "kettle":
                reps = None
                weight = None
                for kcol, w in KETTLE_COLS.items():
                    cv = _num(ws.cell(r, kcol).value)
                    if cv > 0:
                        reps, weight = cv, w
                        break
                if reps is not None:
                    raw[etype] = ("kettle", reps, weight)
            else:
                val = _num(ws.cell(r, rcol).value)
                if val > 0:
                    raw[etype] = val

        rec = {
            "name": str(name).strip(),
            "dob_raw": dob_raw,
            "dob": norm_dob(dob_raw),
            "age": _age_of(dob_raw) if isinstance(dob_raw, (date, datetime)) else 19,
            "row": r,
            "raw": raw,
            "committee_o": committee_o,
            # Агрегаты, подписанные комиссией (verbatim-режим берёт их как есть).
            "committee_agg": {
                STRENGTH: int(_num(ws.cell(r, COL_STRENGTH_O).value)),
                SPEED: int(_num(ws.cell(r, COL_SPEED_O).value)),
                ENDURANCE: int(_num(ws.cell(r, COL_ENDURANCE_O).value)),
                "physical_test_grade": int(_num(ws.cell(r, COL_ITOG100).value)),
            },
        }
        index.add(rec)

    return index


class GradeIndex:
    """Индекс среднего балла.

    by_key  : (norm_name, norm_dob) -> mean_grade(float 0..10)
    by_name : norm_name -> [mean_grade, ...]  (уникальный fallback)
    """

    def __init__(self):
        self.by_key: dict[tuple[str, str], float] = {}
        self.by_name: dict[str, list[float]] = {}
        self.records: list[dict] = []

    def add(self, name, dob_raw, mean_grade: float) -> None:
        self.records.append(
            {"name": str(name).strip(), "dob_raw": dob_raw,
             "dob": norm_dob(dob_raw), "mean_grade": mean_grade}
        )
        self.by_key[_key(name, dob_raw)] = mean_grade
        self.by_name.setdefault(norm_name(name), []).append(mean_grade)

    def lookup(self, name, dob):
        """Вернуть (mean_grade, match_kind) или (None, None)."""
        mg = self.by_key.get(_key(name, dob))
        if mg is not None:
            return mg, "key"
        cands = self.by_name.get(norm_name(name), [])
        if len(cands) == 1:
            return cands[0], "name"
        return None, None


def load_grades(path: Path = GRADE_PATH) -> GradeIndex:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[GRADE_SHEET]

    index = GradeIndex()
    for r in range(2, ws.max_row + 1):  # строка 1 — заголовок
        name = ws.cell(r, GRADE_NAME_COL).value
        if not name:
            continue
        dob_raw = ws.cell(r, GRADE_DOB_COL).value
        mean = ws.cell(r, GRADE_MEAN_COL).value
        if mean in (None, ""):
            continue
        try:
            mean = float(mean)
        except (TypeError, ValueError):
            continue
        index.add(name, dob_raw, mean)

    return index


# ===========================================================================
# Расчёт value_to_store и score для одного упражнения одного человека
# ===========================================================================
def _score_exercise(etype, kind, rcol, ocol, conv, raw_entry, age, ws=None, row=None,
                    committee_o=None):
    """Вернуть (value_to_store, score, extra_params) для одного упражнения.

    raw_entry — то, что положил load_fizo в rec["raw"][etype]:
        число (счётные/бег) либо кортеж ('kettle', reps, weight).
    committee_o — подписанный комиссией балл О (для committee-режима инверсии).

    value_to_store — то, что реально запишется в DAL / протокол (для беговых в
    committee-режиме это примагниченный к сетке порог; выносливость всегда в
    десятичных минутах). score — балл под DAL от value_to_store.
    """
    extra_params: dict = {}

    # --- определяем СЫРОЕ value_to_store для каждого вида упражнения ---
    if kind == "kettle":
        _tag, reps, weight = raw_entry
        raw_value = int(reps)
        extra_params = {"weight": weight}
    elif kind == "int":
        raw_value = int(raw_entry)
    elif kind == "mmss":
        raw_value = mmss_to_decmin(float(raw_entry))  # десятичные минуты
    else:  # 'sec'
        raw_value = float(raw_entry)

    # --- verbatim: храним сырое, балл = дословный О комиссии ---
    if SCORING_MODE == "verbatim":
        score = int(committee_o) if committee_o is not None else conv(
            raw_value, extra_params, age
        )
        return raw_value, score, extra_params

    # --- committee/strict: балл пересчитывается по scoring.py из value ---
    if kind in ("kettle", "int"):
        score = conv(raw_value, extra_params, age)
        return raw_value, score, extra_params

    # Беговые упражнения (kind in {'sec','mmss'}) для committee/strict.
    raw_dv = raw_value

    if SCORING_MODE == "committee":
        table = TIME_TABLES[etype]
        # Инверсия: порог сетки, дающий подписанный О комиссии под DAL.
        if committee_o and committee_o > 0:
            snapped = invert_time(table, committee_o)
        else:
            snapped = None
        if snapped is not None:
            value_to_store = snapped
            score = conv(value_to_store, {}, age)
        else:
            # Фолбэк: О==0 либо не выражается порогом — храним сырое,
            # score = его собственный балл под DAL.
            value_to_store = raw_dv
            score = conv(value_to_store, {}, age)
    else:  # strict
        value_to_store = raw_dv
        score = conv(value_to_store, {}, age)

    return value_to_store, score, extra_params


# ===========================================================================
# Главная функция: compute_fill(name, dob) -> dict
# ===========================================================================
def compute_fill(name, dob, fizo: FizoIndex, grades: GradeIndex) -> dict:
    """Собрать все итоговые значения протокола для одного человека.

    Зеркалит comp_sel_protocol._make_applicant_row / _select_best_exercises:
      * по каждому направлению берём упражнение с МАКСИМАЛЬНЫМ score;
      * strength/speed/endurance_score = max score в направлении (0 если нет);
      * total_score = сумма трёх;
      * physical_test_grade = scoring.compute_global_score(...);
      * mean_grade_scaled = int(mean_grade*10);
      * final_result = physical_test_grade + mean_grade_scaled.
    """
    result = {
        "name": str(name).strip() if name else "",
        "dob": norm_dob(dob),
        "matched": False,
        "match_kind": None,
        "age": None,
        "directions": {STRENGTH: None, SPEED: None, ENDURANCE: None},
        "strength_score": 0,
        "speed_score": 0,
        "endurance_score": 0,
        "total_score": 0,
        "physical_test_grade": 0,
        "mean_grade": None,
        "mean_grade_scaled": None,
        "final_result": 0,
        "exercise_results": [],
    }

    rec, kind = fizo.lookup(name, dob)

    # --- средний балл (сопоставляется независимо от ФИЗО) ---
    mg, _mg_kind = grades.lookup(name, dob)
    if mg is not None:
        result["mean_grade"] = mg
        result["mean_grade_scaled"] = int(mg * 10)

    if rec is None:
        # Нет ФИЗО: физподготовка = 0, но итог включает успеваемость.
        scaled = result["mean_grade_scaled"] or 0
        result["final_result"] = 0 + scaled
        return result

    result["matched"] = True
    result["match_kind"] = kind
    age = rec["age"]
    result["age"] = age

    # Лучшее упражнение в каждом направлении (по МАКСИМАЛЬНОМУ score).
    best: dict[str, dict] = {}
    exercise_results: list[dict] = []

    for (etype, rcol, ocol, conv, direction, number, kind_ex) in EXODEF:
        raw_entry = rec["raw"].get(etype)
        if raw_entry is None:
            continue  # упражнение не выполнялось

        committee_o = int(_num(rec.get("committee_o", {}).get(etype, 0)))

        value_to_store, score, extra_params = _score_exercise(
            etype, kind_ex, rcol, ocol, conv, raw_entry, age,
            committee_o=committee_o,
        )

        exercise_results.append({
            "exercise_type": etype,
            "value_to_store": value_to_store,
            "secondary_score": score,
            "extra_params": extra_params,
        })

        existing = best.get(direction)
        if existing is None or score > existing["score"]:
            best[direction] = {
                "number": EXERCISE_NUMBERS.get(etype, ""),
                "value_to_store": value_to_store,
                "result_display": _format_exercise_value(value_to_store, etype),
                "score": score,
                "exercise_type": etype,
            }

    result["exercise_results"] = exercise_results

    for direction in (STRENGTH, SPEED, ENDURANCE):
        result["directions"][direction] = best.get(direction)

    st = best[STRENGTH]["score"] if STRENGTH in best else 0
    sp = best[SPEED]["score"] if SPEED in best else 0
    en = best[ENDURANCE]["score"] if ENDURANCE in best else 0

    result["strength_score"] = st
    result["speed_score"] = sp
    result["endurance_score"] = en
    result["total_score"] = st + sp + en

    if SCORING_MODE == "verbatim":
        # Итог по 100-балльной — дословно из ведомости комиссии.
        result["physical_test_grade"] = rec["committee_agg"]["physical_test_grade"]
    else:
        result["physical_test_grade"] = scoring.compute_global_score(st, sp, en, age)

    scaled = result["mean_grade_scaled"] or 0
    result["final_result"] = result["physical_test_grade"] + scaled

    return result


def _format_exercise_value(value, exercise_type=None):
    """Зеркало back-end comp_sel_protocol._format_exercise_value (+ мм:сс)."""
    if value is None:
        return ""
    if exercise_type in _MMSS_EXERCISES and isinstance(value, (int, float)):
        total_seconds = round(value * 60)
        return f"{total_seconds // 60}:{total_seconds % 60:02d}"
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return round(value, 2) if isinstance(value, float) else value


# ===========================================================================
# Отчёт: какие имена из ФИЗО НЕ попадут в протокол (raw-индекс наружу)
# ===========================================================================
def unmatched_fizo_report(fizo: FizoIndex, protocol_keys) -> list[dict]:
    """Вернуть список записей ФИЗО, чей (name,dob)-ключ отсутствует среди
    protocol_keys (итерабельное из (name, dob) протоколов). Полезно понять,
    кто из ведомости не найдёт свою строку в протоколах.
    """
    wanted = {_key(n, d) for (n, d) in protocol_keys}
    missing = []
    for rec in fizo.records:
        if _key(rec["name"], rec["dob_raw"]) not in wanted:
            missing.append(rec)
    return missing


# Обратно-совместимый алиас: раньше полный загрузчик назывался load_fizo_full.
# Теперь load_fizo сам читает committee_o, так что это один и тот же вызов.
load_fizo_full = load_fizo


# ===========================================================================
# Демонстрация / самопроверка
# ===========================================================================
def _print_fill(title, fill):
    print(f"\n--- {title} ---")
    print(f"  matched={fill['matched']} match_kind={fill['match_kind']} age={fill['age']}")
    dir_names = {STRENGTH: "СИЛА     ", SPEED: "БЫСТРОТА ", ENDURANCE: "ВЫНОСЛИВ."}
    for d in (STRENGTH, SPEED, ENDURANCE):
        info = fill["directions"][d]
        if info is None:
            print(f"  {dir_names[d]}: —")
        else:
            print(
                f"  {dir_names[d]}: №{info['number']} {info['exercise_type']:18s}"
                f" store={info['value_to_store']!r:>8} disp={info['result_display']!r:>8}"
                f" score={info['score']}"
            )
    print(
        f"  strength={fill['strength_score']} speed={fill['speed_score']}"
        f" endurance={fill['endurance_score']} total={fill['total_score']}"
    )
    print(
        f"  physical_test_grade={fill['physical_test_grade']}"
        f"  mean_grade={fill['mean_grade']} mean_grade_scaled={fill['mean_grade_scaled']}"
        f"  final_result={fill['final_result']}"
    )
    print(f"  exercise_results ({len(fill['exercise_results'])}):")
    for er in fill["exercise_results"]:
        print(f"      {er['exercise_type']:18s} value_to_store={er['value_to_store']!r}"
              f" extra_params={er['extra_params']}")


def main():
    # Самопроверка импорта настоящего scoring.py
    assert hasattr(scoring, "convert_pull_ups"), "scoring.py import failed"
    assert scoring.convert_pull_ups(25, {}, 19) == 100, "scoring sanity check failed"
    assert scoring.convert_run_60m(7.8, {}, 19) == 100, "scoring sanity check failed"
    print("scoring import OK (convert_pull_ups(25)=100, convert_run_60m(7.8)=100)")

    print(f"SCORING_MODE = {SCORING_MODE!r}")

    fizo = load_fizo()
    grades = load_grades()

    print(f"ФИЗО people loaded    : {len(fizo.records)}")
    print(f"grade people loaded   : {len(grades.records)}")

    # Три демо-персоны: 2 из ФИЗО «до 25» + один ВАВТ.
    demos = [
        ("Абрамов Андрей Игоревич", "07.03.2007"),
        ("Авдеев Алексей Дмитриевич", "03.01.2007"),
        ("Архиреев Лев Андреевич", "16.11.2006"),  # ВАВТ (протокол 225 (543))
    ]
    for name, dob in demos:
        fill = compute_fill(name, dob, fizo, grades)
        _print_fill(f"{name} / {dob}", fill)

    print("\nОК: core.py загрузился, посчитал и прошёл самопроверки.")


if __name__ == "__main__":
    main()
