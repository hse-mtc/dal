"""
fill_manual.py — Вариант 1 (ручной): заполняет копии существующих xlsx-протоколов
"Протокол конкурсного отбора" физическими результатами (ФИЗО) и средним баллом.

Логика чисел — ТОЛЬКО из общего модуля core.py (compute_fill). Второй путь
(prod-API) использует тот же core.py, поэтому итоговые значения обязаны совпасть.

Что делает:
  * Копирует каждый протокол из ~/Downloads/dal-fill-db/Протоколы/*.xlsx в
    ~/Downloads/dal-fill-db/artifacts/manual/ (загружает настоящую книгу openpyxl,
    правит на месте, сохраняет по новому пути — формат сохраняется, НЕ регенерирует).
  * По каждому листу идёт по строкам данных (данные с row 20; строка данных — это
    col B = ФИО \n DD.MM.YYYY). Разбирает имя+дату, зовёт core.compute_fill.
  * Пишет колонки по порядку из comp_sel_protocol (1-based A..W):
        G,H,I = STRENGTH  номер,результат,балл
        J,K,L = SPEED
        M,N,O = ENDURANCE
        P     = общий балл (total_score)
        R     = оценка физподготовки (physical_test_grade)
        U     = успеваемость (mean_grade_scaled)
        V     = итоговый результат (final_result)
    Пустое направление -> 3 ячейки пустые. Нет среднего балла -> U не трогаем,
    V = physical_test_grade (зеркало: final = physical_test_grade + (scaled or 0)).
    НЕ трогаем A,B,C,D,E,F,Q,S,T,W.
  * Физику пишем только тем, кто найден в ФИЗО; средний балл — только найденным
    в "Средний балл".
  * Пишет artifacts/manual/report.json и report.txt.

Запуск:  python3 tools/dal-fill/fill_manual.py
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl

import core

# ---------------------------------------------------------------------------
# Пути
# ---------------------------------------------------------------------------
DATA = core.DATA
PROTO_SRC_DIR = DATA / "Протоколы"
OUT_DIR = DATA / "artifacts" / "manual"

DATA_START_ROW = 20  # первая строка данных в каждом листе протокола

# ---------------------------------------------------------------------------
# Колонки протокола (1-based), зеркало comp_sel_protocol._make_applicant_row.
#   Порядок ячеек строки: A(1) № | B(2) ФИО+ДР | C(3) код | D(4) годность |
#   E(5) проф-псих | F(6) возр.группа |
#   G(7)/H(8)/I(9)   СИЛА  номер/результат/балл
#   J(10)/K(11)/L(12) БЫСТРОТА
#   M(13)/N(14)/O(15) ВЫНОСЛИВОСТЬ
#   P(16) общий балл | Q(17) соответствие | R(18) оценка физо |
#   S(19) 10% | T(20) преимущ. | U(21) успеваемость | V(22) итог | W(23) решение
# ---------------------------------------------------------------------------
COL = {
    "ST_NUM": 7,  "ST_RES": 8,  "ST_SCORE": 9,
    "SP_NUM": 10, "SP_RES": 11, "SP_SCORE": 12,
    "EN_NUM": 13, "EN_RES": 14, "EN_SCORE": 15,
    "TOTAL": 16,          # P
    "PHYS_GRADE": 18,     # R
    "MEAN_SCALED": 21,    # U
    "FINAL": 22,          # V
}

DIRECTIONS = [
    (core.STRENGTH, "ST_NUM", "ST_RES", "ST_SCORE"),
    (core.SPEED,    "SP_NUM", "SP_RES", "SP_SCORE"),
    (core.ENDURANCE, "EN_NUM", "EN_RES", "EN_SCORE"),
]


def parse_cell_b(value):
    """Из 'ФИО\\nDD.MM.YYYY' вернуть (name, dob) либо (None, None), если это не
    строка данных (нет перевода строки => заголовок/пусто)."""
    if value is None:
        return None, None
    s = str(value)
    if "\n" not in s:
        return None, None
    parts = s.split("\n")
    name = parts[0].strip()
    dob = parts[1].strip() if len(parts) > 1 else ""
    if not name:
        return None, None
    return name, dob


def fill_sheet(ws, fizo, grades, stats):
    """Заполнить один лист. stats — dict-аккумулятор для отчёта."""
    for r in range(DATA_START_ROW, ws.max_row + 1):
        a = ws.cell(r, 1).value
        # Граница: начало раздела 2 (не прошедшие отбор).
        if isinstance(a, str) and a.strip().startswith("2. Список"):
            break

        name, dob = parse_cell_b(ws.cell(r, 2).value)
        if name is None:
            continue  # спейсер / пустая строка

        stats["total_rows"] += 1
        fill = core.compute_fill(name, dob, fizo, grades)

        # --- Физподготовка: только если найден в ФИЗО ---
        if fill["matched"]:
            stats["physical_filled"] += 1
            for direction, num_k, res_k, score_k in DIRECTIONS:
                info = fill["directions"][direction]
                if info is None:
                    # Направление отсутствует -> оставляем 3 ячейки пустыми.
                    continue
                ws.cell(r, COL[num_k]).value = info["number"]
                ws.cell(r, COL[res_k]).value = info["result_display"]
                ws.cell(r, COL[score_k]).value = info["score"]
            ws.cell(r, COL["TOTAL"]).value = fill["total_score"]
            ws.cell(r, COL["PHYS_GRADE"]).value = fill["physical_test_grade"]
        else:
            stats["unmatched_physical"].append(f"{name} | {dob}")

        # --- Средний балл: только если найден в "Средний балл" ---
        if fill["mean_grade_scaled"] is not None:
            stats["grade_filled"] += 1
            ws.cell(r, COL["MEAN_SCALED"]).value = fill["mean_grade_scaled"]
        else:
            stats["unmatched_grade"] += 1
            stats["unmatched_grade_names"].append(f"{name} | {dob}")

        # --- Итоговый результат V (всегда, зеркало core.final_result) ---
        # final_result = physical_test_grade + (mean_grade_scaled or 0).
        # Если физики нет -> physical_test_grade==0 => V == (scaled or 0).
        ws.cell(r, COL["FINAL"]).value = fill["final_result"]


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    fizo = core.load_fizo()
    grades = core.load_grades()

    src_files = sorted(p for p in PROTO_SRC_DIR.glob("*.xlsx")
                       if not p.name.startswith("~$"))

    report = {
        "scoring_mode": core.SCORING_MODE,
        "fizo_people": len(fizo.records),
        "grade_people": len(grades.records),
        "protocols": [],
        "totals": {
            "total_rows": 0,
            "physical_filled": 0,
            "grade_filled": 0,
            "unmatched_physical": 0,
            "unmatched_grade": 0,
        },
    }

    for src in src_files:
        dst = OUT_DIR / src.name
        shutil.copy2(src, dst)  # сначала байтовая копия (сохраняет всё)

        wb = openpyxl.load_workbook(dst)  # правим настоящую книгу, не регенерируем

        proto_stats = {
            "file": src.name,
            "sheets": [],
            "total_rows": 0,
            "physical_filled": 0,
            "grade_filled": 0,
            "unmatched_physical": [],
            "unmatched_grade": 0,
            "unmatched_grade_names": [],
        }

        for sn in wb.sheetnames:
            ws = wb[sn]
            before = dict(
                total=proto_stats["total_rows"],
                phys=proto_stats["physical_filled"],
                grade=proto_stats["grade_filled"],
            )
            fill_sheet(ws, fizo, grades, proto_stats)
            proto_stats["sheets"].append({
                "sheet": sn,
                "rows": proto_stats["total_rows"] - before["total"],
                "physical_filled": proto_stats["physical_filled"] - before["phys"],
                "grade_filled": proto_stats["grade_filled"] - before["grade"],
            })

        wb.save(dst)

        report["protocols"].append(proto_stats)
        report["totals"]["total_rows"] += proto_stats["total_rows"]
        report["totals"]["physical_filled"] += proto_stats["physical_filled"]
        report["totals"]["grade_filled"] += proto_stats["grade_filled"]
        report["totals"]["unmatched_physical"] += len(proto_stats["unmatched_physical"])
        report["totals"]["unmatched_grade"] += proto_stats["unmatched_grade"]

    # ---- report.json ----
    (OUT_DIR / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # ---- report.txt (человекочитаемый) ----
    lines = []
    lines.append("ОТЧЁТ ЗАПОЛНЕНИЯ ПРОТОКОЛОВ (Вариант 1 — ручной xlsx)")
    lines.append(f"SCORING_MODE = {report['scoring_mode']}")
    lines.append(f"ФИЗО загружено людей : {report['fizo_people']}")
    lines.append(f"Средний балл загружено: {report['grade_people']}")
    lines.append(f"Выходная папка: {OUT_DIR}")
    lines.append("")
    lines.append("ИТОГО ПО ВСЕМ ПРОТОКОЛАМ:")
    t = report["totals"]
    lines.append(f"  строк данных всего     : {t['total_rows']}")
    lines.append(f"  физика заполнена       : {t['physical_filled']}")
    lines.append(f"  средний балл заполнен  : {t['grade_filled']}")
    lines.append(f"  без физики (не найдены): {t['unmatched_physical']}")
    lines.append(f"  без среднего балла     : {t['unmatched_grade']}")
    lines.append("")
    lines.append("=" * 70)

    for p in report["protocols"]:
        lines.append("")
        lines.append(f"ПРОТОКОЛ: {p['file']}")
        lines.append(f"  строк данных всего     : {p['total_rows']}")
        lines.append(f"  физика заполнена       : {p['physical_filled']}")
        lines.append(f"  средний балл заполнен  : {p['grade_filled']}")
        lines.append(f"  без физики (кол-во)    : {len(p['unmatched_physical'])}")
        lines.append(f"  без среднего балла     : {p['unmatched_grade']}")
        for s in p["sheets"]:
            lines.append(
                f"    лист {s['sheet']!r}: строк={s['rows']} "
                f"физика={s['physical_filled']} ср.балл={s['grade_filled']}"
            )
        if p["unmatched_physical"]:
            lines.append("  НЕ НАЙДЕНЫ В ФИЗО:")
            for nm in p["unmatched_physical"]:
                lines.append(f"    - {nm}")

    (OUT_DIR / "report.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")

    print("\n".join(lines))
    print(f"\nreport.json -> {OUT_DIR / 'report.json'}")
    print(f"report.txt  -> {OUT_DIR / 'report.txt'}")


if __name__ == "__main__":
    main()
