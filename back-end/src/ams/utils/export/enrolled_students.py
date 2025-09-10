from pathlib import Path
from tempfile import NamedTemporaryFile
from datetime import date
from typing import Any, Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from lms.models.students import Student
from ams.models.applicants import Applicant

HEADER_GROUPS: List[Tuple[str, List[str]] | str] = [
    "ФГОО ВО в котором обучается студент",
    "ФГОО ВО при которой создан ВУЦ",
    "ВУС",
    "ОВУ",
    "ФГОС",
    "Программа военной подготовки",
    "Год зачисления в ВУЗ",
    "Год начала обучения в ВУЦ",
    "Месяц начала обучения в ВУЦ",
    "Срок обучения (месяцев)",
    "Год окончания ВУЦ",
    "Месяц окончания обучения в ВУЦ",
    "Отметка о завершении военной подготовки",
    "Год окончания ВУЗа",
    "Месяц окончания обучения в ВУЗе",
    "Личный номер",
    "Фамилия",
    "Имя",
    "Отчество",
    "Дата рождения",
    "Место рождения",
    "Националь-ность",
    "Пол",
    "СНИЛС",
    "ИНН",
    (
        "Паспортные данные",
        ["_SER1", "_SER2", "Номер", "Выдан", "Код подразделения", "Кем выдан"],
    ),
    "Адрес регистрации",
    "Номер телефона",
    "Семейное положение",
    "Кол-во детей",
    "Военный комиссариат (по месту жительства)",
    ("Служба ВС РФ (периоды)", ["По призыву", "По контракту"]),
    "Примечание",
    "Статус",
    "Приказ о зачислении",
    "Приказ о отчислении",
    "Причина отчисления",
]


def _flatten_headers(groups):
    cols: List[str] = []
    for item in groups:
        if isinstance(item, str):
            cols.append(item)
        else:
            _, children = item
            cols.extend(children)
    return cols


LEAF_HEADERS: List[str] = _flatten_headers(HEADER_GROUPS)

WIDTHS: Dict[str, float] = {
    "ФГОО ВО в котором обучается студент": 33.0,
    "ФГОО ВО при которой создан ВУЦ": 34.0,
    "ВУС": 10.0,
    "ОВУ": 20.0,
    "ФГОС": 14.0,
    "Программа военной подготовки": 25.0,
    "Год зачисления в ВУЗ": 16.0,
    "Год начала обучения в ВУЦ": 24.0,
    "Месяц начала обучения в ВУЦ": 18.0,
    "Личный номер": 12.0,
    "Фамилия": 18.0,
    "Имя": 16.0,
    "Отчество": 18.0,
    "Дата рождения": 14.0,
    "Место рождения": 28.0,
    "Националь-ность": 14.0,
    "Пол": 10.0,
    "СНИЛС": 16.0,
    "ИНН": 16.0,
    "_SER1": 6.0,
    "_SER2": 6.0,
    "Номер": 14.0,
    "Выдан": 16.0,
    "Код подразделения": 18.0,
    "Кем выдан": 34.0,
    "Адрес регистрации": 30.0,
    "Номер телефона": 16.0,
    "Семейное положение": 18.0,
    "Кол-во детей": 12.0,
    "Военный комиссариат (по месту жительства)": 30.0,
    "По призыву": 18.0,
    "По контракту": 18.0,
    "Примечание": 35.0,
    "Статус": 16.0,
    "Причина отчисления": 28.0,
}
DEFAULT_WIDTH = 12.0

MAIN_HEADER_ROW_HEIGHT = 30.0
SUB_HEADER_ROW_HEIGHT = 38.0
NUMBERS_ROW_HEIGHT = 18.0
DATA_ROW_BASE_HEIGHT = 22.0

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_side = Side(style="thin")
medium_side = Side(style="medium")

all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
all_medium = Border(
    left=medium_side, right=medium_side, top=medium_side, bottom=medium_side
)


def _fmt_date(d: date | None) -> str:
    return d.strftime("%d.%m.%Y") if isinstance(d, date) else ""


def _year(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, date):
        return str(v.year)
    s = str(v).strip()
    return s if s.isdigit() and len(s) in (2, 4) else ""


def _month(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, date):
        return str(v.month)
    try:
        n = int(str(v).strip())
        return str(n) if 1 <= n <= 12 else ""
    except Exception:
        return ""


def _safe_str(v: Any) -> str:
    return "" if v is None else str(v)


def _get_ui(st: Student):
    return getattr(st, "university_info", None)


def _get_bi(st: Student):
    return getattr(st, "birth_info", None)


def _get_ci(st: Student):
    return getattr(st, "contact_info", None)


def _get_pdi(st: Student):
    return getattr(st, "personal_documents_info", None)


def _get_mg(st: Student):
    return getattr(st, "milgroup", None)


def _applicant_mtc_year(st: Student) -> Any:
    ann = getattr(st, "mtc_admission_year", None)
    if ann:
        return ann
    user = getattr(st, "user", None)
    if user:
        app = Applicant.objects.filter(user=user).order_by("-id").first()
        if app and getattr(app, "application_process", None):
            return getattr(app.application_process, "mtc_admission_year", None)
    return None


def _admission_year_like(st: Student) -> str:
    return (
        _year(getattr(st, "mtc_admission_year", None))
        or _year(_applicant_mtc_year(st))
        or _year(getattr(_get_ui(st), "admission_year", None) if _get_ui(st) else None)
        or _year(getattr(_get_ui(st), "enrollment_year", None) if _get_ui(st) else None)
        or _year(getattr(_get_ui(st), "start_year", None) if _get_ui(st) else None)
    )


def _passport_parts(st: Student):
    ps = getattr(st, "passport", None)
    if not ps:
        return "", "", "", "", "", ""
    s1 = s2 = ""
    series = _safe_str(getattr(ps, "series", "")).replace(" ", "")
    if len(series) == 4 and series.isdigit():
        s1, s2 = series[:2], series[2:]
    else:
        raw = _safe_str(getattr(ps, "series", "")).strip()
        if " " in raw:
            parts = [p for p in raw.split(" ") if p]
            if len(parts) >= 2:
                s1, s2 = parts[0], parts[1]
            elif len(parts) == 1:
                s1 = parts[0]
        else:
            s1 = raw
    number = _safe_str(getattr(ps, "code", ""))
    issue_date = _fmt_date(getattr(ps, "issue_date", None))
    dept_code = _safe_str(getattr(ps, "ufms_code", ""))
    issued_by = _safe_str(getattr(ps, "ufms_name", ""))
    return s1, s2, number, issue_date, dept_code, issued_by


def _estimate_height(vals: List[str]) -> float:
    max_lines = 1
    for header, text in zip(LEAF_HEADERS, vals):
        width = WIDTHS.get(header, DEFAULT_WIDTH)
        chars_per_line = max(int(width * 1.15), 3)
        total = 0
        for part in (text or "").split("\n"):
            total += 1 if not part else (len(part) // chars_per_line) + 1
        if total > max_lines:
            max_lines = total
    return max(DATA_ROW_BASE_HEIGHT, DATA_ROW_BASE_HEIGHT * max_lines)


def _outline_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left = medium_side if c == min_col else thin_side
            right = medium_side if c == max_col else thin_side
            top = medium_side if r == min_row else thin_side
            bottom = medium_side if r == max_row else thin_side
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def generate_export(students_qs, _milspecialties_qs) -> Path:
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "НИУ ВШЭ офицеры запаса"

    for i, name in enumerate(LEAF_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(
            name, DEFAULT_WIDTH
        )

    ws.row_dimensions[1].height = MAIN_HEADER_ROW_HEIGHT
    ws.row_dimensions[2].height = SUB_HEADER_ROW_HEIGHT
    ws.row_dimensions[3].height = NUMBERS_ROW_HEIGHT

    main_font = Font(bold=True)
    sub_font = Font(bold=True)
    nums_font = Font(bold=True)

    col = 1
    for item in HEADER_GROUPS:
        if isinstance(item, str):
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            c = ws.cell(row=1, column=col, value=item)
            c.font = main_font
            c.alignment = center
            c.border = all_thin
            col += 1
        else:
            title, children = item
            span = len(children)
            ws.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + span - 1
            )
            c = ws.cell(row=1, column=col, value=title)
            c.font = main_font
            c.alignment = center
            c.border = all_thin

            if title == "Паспортные данные":
                ws.merge_cells(
                    start_row=2, start_column=col, end_row=2, end_column=col + 1
                )
                sc = ws.cell(row=2, column=col, value="Серия")
                sc.font = sub_font
                sc.alignment = center
                sc.border = all_thin
                labels = ["Номер", "Выдан", "Код подразделения", "Кем выдан"]
                for j, lab in enumerate(labels, start=0):
                    sc2 = ws.cell(row=2, column=col + 2 + j, value=lab)
                    sc2.font = sub_font
                    sc2.alignment = center
                    sc2.border = all_thin
            else:
                for j, ch in enumerate(children):
                    sc = ws.cell(row=2, column=col + j, value=ch)
                    sc.font = sub_font
                    sc.alignment = center
                    sc.border = all_thin
            col += span

    for idx in range(1, len(LEAF_HEADERS) + 1):
        nc = ws.cell(row=3, column=idx, value=idx)
        nc.font = nums_font
        nc.alignment = center
        nc.border = all_thin

    ws.freeze_panes = "A4"

    for st in students_qs:
        vals: List[str] = []
        ser1, ser2, pnum, issued, dept, issued_by = _passport_parts(st)
        for header in LEAF_HEADERS:
            if header == "ФГОО ВО в котором обучается студент":
                v = (
                    _safe_str(_get_ui(st).program.faculty.title)
                    if _get_ui(st)
                    and getattr(_get_ui(st), "program", None)
                    and getattr(_get_ui(st).program, "faculty", None)
                    else ""
                )
            elif header == "ФГОО ВО при которой создан ВУЦ":
                v = (
                    _safe_str(_get_mg(st).milfaculty.title)
                    if _get_mg(st) and getattr(_get_mg(st), "milfaculty", None)
                    else ""
                )
            elif header == "ВУС":
                v = (
                    _safe_str(_get_mg(st).milspecialty.code)
                    if _get_mg(st) and getattr(_get_mg(st), "milspecialty", None)
                    else ""
                )
            elif header == "ОВУ":
                v = _safe_str(_get_mg(st).title) if _get_mg(st) else ""
            elif header == "ФГОС":
                v = (
                    _safe_str(_get_ui(st).program.code)
                    if _get_ui(st) and getattr(_get_ui(st), "program", None)
                    else ""
                )
            elif header == "Программа военной подготовки":
                v = (
                    _safe_str(_get_mg(st).milspecialty.title)
                    if _get_mg(st) and getattr(_get_mg(st), "milspecialty", None)
                    else ""
                )
            elif header == "Год зачисления в ВУЗ":
                v = _admission_year_like(st)
            elif header == "Год начала обучения в ВУЦ":
                v = _admission_year_like(st)
            elif header == "Месяц начала обучения в ВУЦ":
                v = "9"
            elif header == "Срок обучения (месяцев)":
                v = ""
            elif header == "Год окончания ВУЦ":
                v = ""
            elif header == "Месяц окончания обучения в ВУЦ":
                v = ""
            elif header == "Отметка о завершении военной подготовки":
                v = ""
            elif header == "Год окончания ВУЗа":
                v = _year(
                    getattr(_get_ui(st), "graduation_date", None)
                    if _get_ui(st)
                    else None
                ) or _year(
                    getattr(_get_ui(st), "graduation_year", None)
                    if _get_ui(st)
                    else None
                )
            elif header == "Месяц окончания обучения в ВУЗе":
                v = _month(
                    getattr(_get_ui(st), "graduation_date", None)
                    if _get_ui(st)
                    else None
                )
            elif header == "Личный номер":
                v = _safe_str(st.id)
            elif header == "Фамилия":
                v = _safe_str(st.surname)
            elif header == "Имя":
                v = _safe_str(st.name)
            elif header == "Отчество":
                v = _safe_str(getattr(st, "patronymic", "") or "")
            elif header == "Дата рождения":
                v = _fmt_date(
                    getattr(_get_bi(st), "date", None) if _get_bi(st) else None
                )
            elif header == "Место рождения":
                v = _safe_str(getattr(_get_bi(st), "place", "") if _get_bi(st) else "")
            elif header == "Националь-ность":
                v = _safe_str(getattr(st, "citizenship", "") or "")
            elif header == "Пол":
                v = (
                    "мужской"
                    if (
                        _safe_str(getattr(st, "gender", "")).lower()
                        in ("m", "male", "м", "муж", "мужской")
                    )
                    else (
                        "женский"
                        if (
                            _safe_str(getattr(st, "gender", "")).lower()
                            in ("f", "female", "ж", "жен", "женский")
                        )
                        else "мужской"
                    )
                )
            elif header == "СНИЛС":
                v = _safe_str(
                    getattr(_get_pdi(st), "insurance_number", "")
                    if _get_pdi(st)
                    else ""
                )
            elif header == "ИНН":
                v = _safe_str(
                    getattr(_get_pdi(st), "tax_id", "") if _get_pdi(st) else ""
                )
            elif header == "_SER1":
                v = ser1
            elif header == "_SER2":
                v = ser2
            elif header == "Номер":
                v = pnum
            elif header == "Выдан":
                v = issued
            elif header == "Код подразделения":
                v = dept
            elif header == "Кем выдан":
                v = issued_by
            elif header == "Адрес регистрации":
                v = _safe_str(getattr(st, "permanent_address", "") or "")
            elif header == "Номер телефона":
                v = _safe_str(
                    getattr(_get_ci(st), "personal_phone_number", "")
                    if _get_ci(st)
                    else ""
                )
            elif header == "Семейное положение":
                v = ""
            elif header == "Кол-во детей":
                v = ""
            elif header == "Военный комиссариат (по месту жительства)":
                v = _safe_str(getattr(st, "recruitment_office", "") or "")
            elif header == "По призыву":
                v = ""
            elif header == "По контракту":
                v = ""
            elif header == "Примечание":
                v = ""
            elif header == "Статус":
                v = (
                    _safe_str(st.get_status_display())
                    if hasattr(st, "get_status_display")
                    else _safe_str(getattr(st, "status", ""))
                )
            elif header == "Приказ о зачислении":
                v = ""
            elif header == "Приказ о отчислении":
                v = ""
            elif header == "Причина отчисления":
                v = ""
            else:
                v = ""
            vals.append(v)

        row_idx = ws.max_row + 1
        ws.row_dimensions[row_idx].height = _estimate_height(vals)
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = center
            c.border = all_thin

    last_row = ws.max_row
    last_col = len(LEAF_HEADERS)
    _outline_range(ws, 1, 2, 1, last_col)
    _outline_range(ws, 3, 3, 1, last_col)
    if last_row >= 4:
        _outline_range(ws, 4, last_row, 1, last_col)

    for col_idx in range(1, last_col + 1):
        ws.cell(row=1, column=col_idx).font = main_font
        if ws.cell(row=2, column=col_idx).value:
            ws.cell(row=2, column=col_idx).font = sub_font
        ws.cell(row=3, column=col_idx).font = nums_font
        ws.cell(row=1, column=col_idx).alignment = center
        ws.cell(row=2, column=col_idx).alignment = center
        ws.cell(row=3, column=col_idx).alignment = center

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return Path(tmp.name)
