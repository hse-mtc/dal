import re
import datetime

import openpyxl
from openpyxl.reader.excel import load_workbook

MONTH_TO_NUM = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}

MAP_LESSONS_TO_ORDINAL = {"1-2": 1, "3-4": 2, "5-6": 3}


def process_date(date):
    cur_year = datetime.datetime.now().year
    while "  " in date:
        date = date.replace("  ", " ")
    day, month = date.split()
    date = f"{cur_year}-{MONTH_TO_NUM[month]:02d}-{int(day):02d}"
    return date


def process_field(text):
    text = text.replace("\n", " ")
    while "  " in text:
        text = text.replace("  ", " ")

    teachers_regexp = re.compile(
        r"(?P<surname>[А-ЯЁ][а-яё]+) ?"
        r"(?P<name>[А-ЯЁ])[\.,] ?"
        r"((?P<patronymic>[А-ЯЁ])[\.,]?)?"
    )
    teachers = re.finditer(teachers_regexp, text)
    teachers = list(
        map(
            lambda x: (
                x.group("surname").lower().strip(),
                x.group("name").lower().strip(),
            ),
            teachers,
        )
    )

    if teachers:
        teachers_start = re.search(teachers_regexp, text).start()
    else:
        teachers_start = None
    classroom = re.search(r"[Аа]уд\.? ?((\d,? ?)*)", text)
    if not classroom:
        if "плац" in text.lower():
            classroom = ["плац"]
            classroom_start = text.lower().find("плац")
        else:
            classroom = []
            classroom_start = None
    else:
        classroom_start = classroom.start()
        classroom = re.findall(r"\d+", classroom.group())
        classroom = list(map(str, classroom))

    if classroom_start is not None:
        subj = text[:classroom_start].strip()
    else:
        if teachers_start is not None:
            subj = text[:teachers_start].strip()
        else:
            subj = text.strip()
    return {"subject": subj, "classrooms": classroom, "teachers": teachers}


def parse_timetable(timetable_path: str):
    wb: openpyxl.workbook.Workbook = load_workbook(timetable_path)
    ws = wb.active
    x = 0
    msls = list(ws.merged_cells)
    for merged_cell in msls:
        x += 1
        top_left_cell = ws[merged_cell.coord][0][0]
        val = top_left_cell.value
        ws.unmerge_cells(
            start_row=merged_cell.min_row,
            start_column=merged_cell.min_col,
            end_row=merged_cell.max_row,
            end_column=merged_cell.max_col,
        )
        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
            for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = val
    headers = {}
    date_header_index = {}
    row_header = {}
    for i in range(5, ws.max_column + 1):
        headers[i] = {}
        headers[i]["week"] = ws.cell(row=1, column=i).value
        headers[i]["ordinal"] = ws.cell(row=3, column=i).value
    now_ind = 2
    good_rows = []
    for i in range(4, ws.max_row + 1):
        if ws.cell(row=i, column=1).value:
            row_header[i] = {}
            row_header[i]["week_day"] = ws.cell(row=i, column=1).value
            row_header[i]["milspec"] = ws.cell(row=i, column=2).value
            row_header[i]["milgroup"] = ws.cell(row=i, column=3).value
            date_header_index[i] = now_ind
            good_rows.append(i)
        else:
            now_ind = i
    final = []
    for i in range(5, ws.max_column + 1):
        for elem in good_rows:
            date = ws.cell(column=i, row=date_header_index[elem]).value
            ss = ws.cell(row=elem, column=i).value
            if ss:
                res = {
                    **row_header[elem],
                    **headers[i],
                    **process_field(ss),
                    "date": process_date(date),
                }
                res["milgroup"] = str(res["milgroup"])
                if res["ordinal"] in MAP_LESSONS_TO_ORDINAL:
                    res["ordinal"] = MAP_LESSONS_TO_ORDINAL[res["ordinal"]]
                    final.append(res)
    return final
