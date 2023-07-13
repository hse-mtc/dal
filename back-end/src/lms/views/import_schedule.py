from drf_spectacular.utils import extend_schema
from rest_framework import status, generics
from rest_framework.request import Request
from rest_framework.response import Response

from auth.models import Permission
from auth.permissions import BasePermission
from common.models.subjects import Subject
from common.parsers import MultiPartWithJSONParser
import datetime
from lms.models.common import Milgroup

import openpyxl
from openpyxl import load_workbook
import re

from lms.models.lessons import Room, Lesson
from lms.models.teachers import Teacher
from lms.serializers.import_schedule import (
    ImportParsedSerializer,
    ParseScheduleSerializer,
)
from lms.serializers.lessons import LessonParsedSerializer
from lms.views.lessons import LessonPermission

month_to_num = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}

MAP_LESSONS_TO_ORDINAL = {"1-2": 1, "3-4": 2, "5-6": 3}


def process_date(date):
    cur_year = datetime.datetime.now().year
    while "  " in date:
        date = date.replace("  ", " ")
    day, month = date.split()
    date = f"{cur_year}-{month_to_num[month]:02d}-{int(day):02d}"
    return date


def process_field(text):
    text = text.replace("\n", " ")
    while "  " in text:
        text = text.replace("  ", " ")

    # text = text.strip().split(" ")
    teachers_regexp = re.compile(
        r"(?P<surname>[А-ЯЁ][а-яё]+) ?"
        r"(?P<name>[А-ЯЁ])[\.,] ?"
        r"((?P<patronymic>[А-ЯЁ])[\.,]?)?"
    )
    teachers = re.finditer(teachers_regexp, text)
    teachers = list(
        map(
            lambda x: (
                x.group("surname").lower().strip(),
                x.group("name").lower().strip(),
            ),
            teachers,
        )
    )

    if teachers:
        teachers_start = re.search(teachers_regexp, text).start()
    else:
        teachers_start = None
    classroom = re.search(r"[Аа]уд\.? ?((\d,? ?)*)", text)
    if not classroom:
        if "плац" in text.lower():
            classroom = ["плац"]
            classroom_start = text.lower().find("плац")
        else:
            classroom = []
            classroom_start = None
    else:
        classroom_start = classroom.start()
        classroom = re.findall(r"\d+", classroom.group())
        classroom = list(map(str, classroom))

    if classroom_start is not None:
        subj = text[:classroom_start].strip()
    else:
        if teachers_start is not None:
            subj = text[:teachers_start].strip()
        else:
            subj = text.strip()
    return {"subject": subj, "classrooms": classroom, "teachers": teachers}


def guess_subject_name(subject_name):
    if "ТП" in subject_name:
        return "Тактическая подготовка"
    if "ТСП" in subject_name:
        return "Тактико-специальная подготовка"
    if "ВСП" in subject_name:
        return "Военно-специальная подготовка"
    if "ВПП" in subject_name:
        return "Военно-политическая подготовка"
    return ""


def guess_subject_type(subject_name):
    if " Л " in subject_name:
        return Lesson.Type.LECTURE
    if " С " in subject_name:
        return Lesson.Type.SEMINAR
    if " ГЗ " in subject_name:
        return Lesson.Type.GROUP
    if " ПЗ " in subject_name:
        return Lesson.Type.PRACTICE
    if "зачет" in subject_name.lower().replace("ё", "е"):
        return Lesson.Type.FINAL_TEST
    if "экзамен" in subject_name.lower():
        return Lesson.Type.EXAM
    return Lesson.Type.OTHER


def parse_timetable(timetable_path: str):
    wb: openpyxl.workbook.Workbook = load_workbook(timetable_path)
    ws = wb.active
    x = 0
    msls = list(ws.merged_cells)
    for merged_cell in msls:
        x += 1
        top_left_cell = ws[merged_cell.coord][0][0]
        val = top_left_cell.value
        ws.unmerge_cells(
            start_row=merged_cell.min_row,
            start_column=merged_cell.min_col,
            end_row=merged_cell.max_row,
            end_column=merged_cell.max_col,
        )
        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
            for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = val
    headers = {}
    date_header_index = {}
    row_header = {}
    for i in range(5, ws.max_column + 1):
        headers[i] = {}
        headers[i]["week"] = ws.cell(row=1, column=i).value
        headers[i]["ordinal"] = ws.cell(row=3, column=i).value
    now_ind = 2
    good_rows = []
    for i in range(4, ws.max_row + 1):
        if ws.cell(row=i, column=1).value:
            row_header[i] = {}
            row_header[i]["week_day"] = ws.cell(row=i, column=1).value
            row_header[i]["milspec"] = ws.cell(row=i, column=2).value
            row_header[i]["milgroup"] = ws.cell(row=i, column=3).value
            date_header_index[i] = now_ind
            good_rows.append(i)
        else:
            now_ind = i
    final = []
    for i in range(5, ws.max_column + 1):
        for elem in good_rows:
            date = ws.cell(column=i, row=date_header_index[elem]).value
            ss = ws.cell(row=elem, column=i).value
            if ss:
                res = {
                    **row_header[elem],
                    **headers[i],
                    **process_field(ss),
                    "date": process_date(date),
                }
                res["milgroup"] = str(res["milgroup"])
                if res["ordinal"] in MAP_LESSONS_TO_ORDINAL:
                    res["ordinal"] = MAP_LESSONS_TO_ORDINAL[res["ordinal"]]
                    final.append(res)
    return final


class ImportSchedulePermission(BasePermission):
    permission_class = "import-permission"
    view_name_rus = "Импорт расписания из файла"
    scopes = [
        Permission.Scope.ALL,
    ]


@extend_schema(tags=["import-schedule"])
class ParseScheduleView(generics.GenericAPIView):
    parser_classes = [MultiPartWithJSONParser]
    permission_classes = [LessonPermission]
    serializer_class = ParseScheduleSerializer

    def post(self, request: Request):
        if "content" not in request.data:
            return Response(
                {"detail": "Bad request: no file passed"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        file = request.data["content"]

        lessons_list = parse_timetable(file)
        milgroups = set()
        teachers = set()
        classrooms = set()

        for elem in lessons_list:
            milgroups.add(elem["milgroup"])
            for t in elem["teachers"]:
                teachers.add(t)
            for c in elem["classrooms"]:
                classrooms.add(c)

        milgroups_cache = {}
        classrooms_cache = {}
        teachers_cache = {}

        for milgroup in milgroups:
            milg_filtered = Milgroup.objects.filter(title__iexact=milgroup)
            if milg_filtered:
                milgroups_cache[milgroup.lower()] = milg_filtered[0]
        for classroom in classrooms:
            classroom_filtered = Room.objects.filter(title__iexact=classroom)
            if classroom_filtered:
                classrooms_cache[classroom.lower()] = classroom_filtered[0]

        for teacher in teachers:
            surname, name = teacher
            teacher_filtered = Teacher.objects.filter(
                surname__iexact=surname, name__istartswith=name[0]
            )

            if teacher_filtered:
                teachers_cache[teacher] = teacher_filtered[0]
        parsed_lessons_list = []
        for elem in lessons_list:
            lesson = {"input": {}, "parsed": {}}
            lesson_name_input = elem["subject"]
            lesson["input"]["lesson_name_input"] = lesson_name_input

            lesson_subject_title = guess_subject_name(lesson_name_input)
            lesson["input"]["lesson_subject_title"] = lesson_subject_title

            lesson_type = guess_subject_type(lesson_name_input)
            lesson["parsed"]["type"] = lesson_type

            lesson["parsed"]["type"] = lesson_type
            lesson["parsed"]["date"] = elem["date"]
            lesson["parsed"]["ordinal"] = elem["ordinal"]
            if elem["milgroup"] in milgroups_cache:
                milgroup = milgroups_cache[elem["milgroup"]]
                lesson["parsed"]["milgroup_pk"] = milgroup.pk
                lesson["parsed"]["milgroup_title"] = milgroup.title
                subjects = Subject.objects.filter(
                    milspecialty=milgroup.milspecialty, title=lesson_subject_title
                )

                if len(subjects) > 0:
                    lesson_subject = subjects[0]
                    lesson["parsed"]["subject_pk"] = lesson_subject.pk
                    lesson["parsed"]["subject_title"] = lesson_subject.pk
                else:
                    lesson["parsed"]["subject_pk"] = None
                    lesson["parsed"]["subject_title"] = None
            else:
                lesson["parsed"]["milgroup_pk"] = None
                lesson["parsed"]["milgroup_title"] = None
                lesson["parsed"]["subject_pk"] = None
                lesson["parsed"]["subject_title"] = None

            teacher_names = elem["teachers"]
            lesson["input"]["teachers"] = teacher_names
            if len(teacher_names) > 0 and teacher_names[0] in teachers_cache:
                teacher = teachers_cache[teacher_names[0]]
                lesson["parsed"]["teacher_pk"] = teacher.pk
                lesson["parsed"][
                    "teacher_name"
                ] = f"{teacher.surname} {teacher.name} {teacher.patronymic}"
            else:
                lesson["parsed"]["teacher_pk"] = None
                lesson["parsed"]["teacher_name"] = None

            classrooms = elem["classrooms"]
            if len(classrooms) > 0 and classrooms[0] in classrooms_cache:
                classroom = classrooms_cache[classrooms[0]]
                lesson["parsed"]["classroom_pk"] = classroom.pk
                lesson["parsed"]["classroom_title"] = classroom.title
            else:
                lesson["parsed"]["classroom_pk"] = None
                lesson["parsed"]["classroom_title"] = None

            parsed_lessons_list.append(lesson)
        return Response(parsed_lessons_list)


@extend_schema(tags=["import-schedule"])
class ImportParsedView(generics.GenericAPIView):
    permission_classes = [ImportSchedulePermission]
    serializer_class = ImportParsedSerializer

    def post(self, request: Request):
        valid_data = []
        if "parsed" not in request.data:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        for data in request.data["parsed"]:
            serializer = LessonParsedSerializer(data=data)
            if serializer.is_valid():
                valid_data.append(data)
        request.data["parsed"] = valid_data
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"created": len(valid_data)}, status=status.HTTP_201_CREATED)
